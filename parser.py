"""
Parse Active Visa Matters and Capacity Planning xlsx files,
merge data per candidate into hydrated Candidate objects.
"""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl

from config import SA_CREDENTIAL_COLUMNS, SKIP_SHEETS
from models import Candidate, Client
from config import detect_parent_company, detect_branch_name

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _str(val: Any) -> str:
    """Convert a cell value to a stripped string, or empty string if None."""
    if val is None:
        return ""
    return str(val).strip()


def _parse_date(val: Any) -> Optional[date]:
    """Parse a cell value into a date. Handles datetime objects and common string formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    # Try common formats
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    logger.warning(f"Could not parse date: {s!r}")
    return None


def _parse_float(val: Any) -> Optional[float]:
    """Parse a cell value into a float, or None."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _valid_file_name(val: Any) -> Optional[str]:
    """Return a valid File_Name string, or None if the row should be skipped."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if val == 0.0:
            return None
        return str(int(val)) if val == int(val) else str(val)
    s = str(val).strip()
    return s if s else None


def _build_col_map(header_row: tuple) -> dict[str, int]:
    """Build a column-name-to-index map from a header row.

    Stores exact-case keys. _get() falls back to case-insensitive lookup
    so minor header casing differences in spreadsheets don't silently break parsing.
    """
    col_map = {}
    for i, cell in enumerate(header_row):
        if cell is not None:
            col_map[str(cell).strip()] = i
    return col_map


def _get(row: tuple, col_map: dict, col_name: str, default: Any = None) -> Any:
    """Get a value from a row by column name. Falls back to case-insensitive match."""
    idx = col_map.get(col_name)
    if idx is None:
        # Case-insensitive fallback
        col_name_lower = col_name.lower()
        for key, i in col_map.items():
            if key.lower() == col_name_lower:
                idx = i
                break
    if idx is None or idx >= len(row):
        return default
    return row[idx]


# ---------------------------------------------------------------------------
# Sheet Parsers
# ---------------------------------------------------------------------------

def _parse_candidate_sheet(ws, sheet_source: str) -> dict[str, Candidate]:
    """Parse Preparing, Processing, Finalised, or PausedFuture sheet."""
    candidates = {}
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return candidates

    col_map = _build_col_map(rows[0])
    if not col_map:
        logger.warning(f"Sheet '{ws.title}' has an empty or unreadable header row — skipping")

    for row in rows[1:]:
        fn = _valid_file_name(_get(row, col_map, "File_Name"))
        if not fn:
            continue

        c = Candidate(
            file_name=fn,
            given_names=_str(_get(row, col_map, "Given_Names")),
            family_names=_str(_get(row, col_map, "Family_Names")),
            company=_str(_get(row, col_map, "Company")),
            visa_type=_str(_get(row, col_map, "Type")),
            sheet_source=sheet_source,
            nomination_status=_str(_get(row, col_map, "Main_Nomination_Status")),
            visa_status=_str(_get(row, col_map, "Main_Visa_Status")),
            other_app_status=_str(_get(row, col_map, "Other_App_Status")),
            sa_status=_str(_get(row, col_map, "SA_Status")) or None,
            notes=_str(_get(row, col_map, "Notes")),
            extra_notes=_str(_get(row, col_map, "Extra Notes")),
            occupation=_str(_get(row, col_map, "Occupation")),
        )

        # Date fields differ between sheets
        if sheet_source == "preparing":
            c.visa_lodged_date = _parse_date(_get(row, col_map, "Visa Lodged"))
        else:
            c.visa_lodged_date = _parse_date(_get(row, col_map, "Date Visa Lodged"))

        if sheet_source == "finalised":
            c.visa_finalised_date = _parse_date(_get(row, col_map, "Date Visa Finalised"))

        # Handover field (Preparing only)
        handover = _str(_get(row, col_map, "Handover"))
        if handover:
            c.extra_notes = f"{c.extra_notes} | Handover: {handover}".strip(" |")

        candidates[fn] = c

    return candidates


def _parse_sa_sheet(ws) -> dict[str, dict]:
    """Parse Skills Assessments or S.A. Completed sheet. Returns dict keyed by File_Name."""
    sa_data = {}
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return sa_data

    col_map = _build_col_map(rows[0])

    for row in rows[1:]:
        fn = _valid_file_name(_get(row, col_map, "File_Name"))
        if not fn:
            continue

        raw_progress = _parse_float(_get(row, col_map, "Progress"))
        sa_progress = max(0.0, min(1.0, raw_progress)) if raw_progress is not None else None
        sa_data[fn] = {
            "sa_progress": sa_progress,
            "sa_notes": _str(_get(row, col_map, "Notes")),
            "occupation": _str(_get(row, col_map, "Occupation")),
            "acts": _parse_float(_get(row, col_map, "ACTS")),
            "commencement_date": _parse_date(_get(row, col_map, "Commencement Date")),
            "tsa": _str(_get(row, col_map, "TSA")),
        }

    return sa_data


def _parse_cp_team_sheet(ws) -> dict[str, dict]:
    """Parse a Capacity Planning team sheet (Deisy/Erika/Gabriel).
    Returns dict keyed by File_Name with aggregated tasks and notes."""
    cp_data: dict[str, dict] = {}
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return cp_data

    col_map = _build_col_map(rows[0])

    for row in rows[1:]:
        fn = _valid_file_name(_get(row, col_map, "File_Name"))
        if not fn:
            continue

        task = _str(_get(row, col_map, "Task"))
        note = _str(_get(row, col_map, "Notes"))

        if fn not in cp_data:
            cp_data[fn] = {"tasks": [], "notes": []}

        if task:
            cp_data[fn]["tasks"].append(task)
        if note:
            cp_data[fn]["notes"].append(note)

    return cp_data


def _parse_lodgements_sheet(ws) -> dict[str, dict]:
    """Parse Lodgements sheet. Returns dict keyed by File No."""
    lodgement_data = {}
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return lodgement_data

    col_map = _build_col_map(rows[0])

    for row in rows[1:]:
        fn = _valid_file_name(_get(row, col_map, "File No"))
        if not fn:
            continue

        lodgement_data[fn] = {
            "status": _str(_get(row, col_map, "Status")),
            "visa_subclass": _str(_get(row, col_map, "Visa Subclass")),
            "comments": _str(_get(row, col_map, "Comments")),
        }

    return lodgement_data


# ---------------------------------------------------------------------------
# Main Parse + Merge
# ---------------------------------------------------------------------------

def parse_workbooks(visa_matters_path: str | Path,
                    capacity_planning_path: str | Path) -> list[Candidate]:
    """Parse both xlsx workbooks and return a merged list of Candidate objects."""

    vm_wb = openpyxl.load_workbook(str(visa_matters_path), read_only=True, data_only=True)
    cp_wb = openpyxl.load_workbook(str(capacity_planning_path), read_only=True, data_only=True)

    # --- Active Visa Matters ---
    all_candidates: dict[str, Candidate] = {}

    sheet_map = {
        "Preparing": "preparing",
        "Processing": "processing",
        "Finalised": "finalised",
        "PausedFuture": "paused",
    }

    for sheet_name, source in sheet_map.items():
        if sheet_name in vm_wb.sheetnames:
            candidates = _parse_candidate_sheet(vm_wb[sheet_name], source)
            all_candidates.update(candidates)
            logger.info(f"Parsed {len(candidates)} candidates from {sheet_name}")
        else:
            logger.warning(f"Sheet '{sheet_name}' not found in Active Visa Matters")

    # --- Skills Assessments ---
    sa_data: dict[str, dict] = {}
    for sa_sheet_name in ("Skills Assessments", "S.A. Completed"):
        if sa_sheet_name in vm_wb.sheetnames:
            sheet_sa = _parse_sa_sheet(vm_wb[sa_sheet_name])
            sa_data.update(sheet_sa)
            logger.info(f"Parsed {len(sheet_sa)} SA records from {sa_sheet_name}")

    # --- Capacity Planning team sheets ---
    cp_data: dict[str, dict] = {}
    for team_sheet in ("Deisy", "Erika", "Gabriel"):
        if team_sheet in cp_wb.sheetnames:
            sheet_cp = _parse_cp_team_sheet(cp_wb[team_sheet])
            # Merge into aggregate (a candidate may appear on multiple team sheets)
            for fn, data in sheet_cp.items():
                if fn not in cp_data:
                    cp_data[fn] = {"tasks": [], "notes": []}
                cp_data[fn]["tasks"].extend(data["tasks"])
                cp_data[fn]["notes"].extend(data["notes"])
            logger.info(f"Parsed CP data from {team_sheet}")

    # --- Lodgements ---
    lodgement_data: dict[str, dict] = {}
    if "Lodgements" in cp_wb.sheetnames:
        lodgement_data = _parse_lodgements_sheet(cp_wb["Lodgements"])
        logger.info(f"Parsed {len(lodgement_data)} lodgement records")

    # --- Close workbooks ---
    vm_wb.close()
    cp_wb.close()

    # --- Merge pass ---
    for fn, candidate in all_candidates.items():
        # Attach SA data
        if fn in sa_data:
            sa = sa_data[fn]
            candidate.sa_progress = sa.get("sa_progress")
            candidate.sa_notes = sa.get("sa_notes", "")
            if sa.get("occupation") and not candidate.occupation:
                candidate.occupation = sa["occupation"]

        # Attach CP data
        if fn in cp_data:
            candidate.recent_tasks = cp_data[fn]["tasks"]
            candidate.task_notes = cp_data[fn]["notes"]

        # Attach Lodgement data
        if fn in lodgement_data:
            lodg = lodgement_data[fn]
            candidate.lodgement_status = lodg.get("status")
            candidate.lodgement_comments = lodg.get("comments")

    return list(all_candidates.values())


def group_by_client(candidates: list[Candidate]) -> dict[str, Client]:
    """Group candidates into Client objects by company, with multi-branch detection."""
    client_map: dict[str, Client] = {}

    for c in candidates:
        company = c.company.strip()
        if not company:
            continue

        if company not in client_map:
            parent = detect_parent_company(company)
            branch = detect_branch_name(company, parent) if parent else None
            client_map[company] = Client(
                company_name=company,
                parent_company=parent,
                branch=branch,
            )
        client_map[company].candidates.append(c)

    return client_map
